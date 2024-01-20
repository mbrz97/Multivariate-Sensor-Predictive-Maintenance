import math
import os
import re
from settings import ROOT_PATH
from datetime import datetime, timedelta
import pandas as pd
from datetime import datetime
import xlrd

field_pattern = r'name1 | name2 | name3 | name4 | name5 | name6'

content_dict = {}

files_w_one_column = []
files_w_no_pd = []
none_files = []


# mkey = 'nosrat_nth-1'
# for i in range(len(content_dict[mkey])):
#     if type(content_dict[mkey][i][0]) == str:
#         print('Gap between: 'content_dict[mkey][i-1][0], ', and :'content_dict[mkey][i+1][0])


# def list_of_lists_to_excel(data, filename):
#     import pandas as pd
#     df = pd.DataFrame(data)  # Convert the list of lists to a DataFrame
#     df.to_excel(filename, index=False)  # Write the DataFrame to an Excel file
#
# # Example data: list of lists
# data = [
#     ['Name', 'Age', 'City'],
#     ['John', 25, 'New York'],
#     ['Alice', 30, 'Los Angeles'],
#     ['Bob', 35, 'Chicago']
# ]
# output_name =
#
# # Call the function with the data and desired filename
# list_of_lists_to_excel(data, 'output.xlsx')


def get_field_well(file_name):
    well_patterns = r'[A-Za-z]+[0-9]*[A-Za-z]*-[A-Za-z]*[0-9]+ | [A-Za-z]+-[0-9]+[A-Za-z]+ | [A-Za-z]+[0-9]+'
    file_name = file_name.lower()
    assert (file_name.find('.csv') >= 0 or file_name.find('.xlsx') >= 0)

    field_name = re.search(pattern=field_pattern, string=file_name)
    well_name = re.search(pattern=well_patterns, string=file_name)

    if field_name:
        field_name = field_name.group()
        field_name = field_name.strip()
    if well_name:
        well_name = well_name.group()
        well_name = well_name.strip()

    return field_name, well_name


# def check_col_fix(file_name):
#     import pandas as pd
#     try:
#         if file_name.find('.xlsx') >= 0:
#             data = pd.read_excel(file_name, names=['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'])
#         elif file_name.find('.csv') >= 0:
#             data = pd.read_csv(file_name, names=['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'])
#             return
#         else:
#             return None
#         return None
#     except ValueError:
#         return file_name


def read_file_contents(file_name):
    # print(file_name)
    import pandas as pd

    no_pd = False

    try:
        if file_name.find('.xlsx') >= 0:
            data = pd.read_excel(file_name, names=['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'])
        elif file_name.find('.csv') >= 0:
            data = pd.read_csv(file_name, names=['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'])
        else:
            return None
    except ValueError:
        try:
            if file_name.find('.xlsx') >= 0:
                data = pd.read_excel(file_name, names=['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'])
                no_pd = True
            elif file_name.find('.csv') >= 0:
                data = pd.read_csv(file_name, names=['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'])
                no_pd = True
            else:
                return None
        except ValueError:
            if file_name.find('.xlsx') >= 0:
                data = pd.read_excel(file_name)
            elif file_name.find('.csv') >= 0:
                data = pd.read_csv(file_name)
            num_columns = data.shape[1]
            coll = data.columns.values
            if num_columns == 1:
                # files_w_one_column.append(file_name)
                # pd.set_option('display.max_rows', None)  # Show all rows
                # pd.set_option('display.max_columns', None)  # Show all columns
                # pd.set_option('display.width', None)  # Disable column width restriction
                # pd.set_option('display.max_colwidth', None)  # Disable column width restriction for cell contents
                #
                # print(data.iloc[:,0])
                data = data.iloc[:, 0].str.split(',', n=9, expand=True)
                data = data.iloc[:, 0:9]
                data.columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
                # print(data)
            else:
                return None

        # lst = ['Date/Time', 'pi', 'pd', 'ti', 'tm', 'vx', 'vz', 'vt', 'ct']
        #
        # if file_name.find('.xlsx') >= 0:
        #     data = pd.read_excel(file_name)
        # elif file_name.find('.csv') >= 0:
        #     data = pd.read_csv(file_name)
        # else:
        #     return None
        # data_head = data.head(10)
        #
        # for i, j in enumerate(lst):
        #     is_match = data_head.isin(lst[i]).all(axis=1).any()
        #     if is_match:
        #         continue
        #     else:
        #         val_append = data.loc[i]
        #         data_appended = data.append(val_append, ignore_index=True)
        #         data = data_appended

    # first_row = True

    if no_pd:
        files_w_no_pd.append(file_name)
        data.insert(2, 'pd', 0)
        new_col_names = {'A': 'A', 'B': 'B', 'pd': 'C', 'C': 'D', 'D': 'E', 'E': 'F', 'F': 'G', 'G': 'H', 'H': 'I'}
        data = data.rename(columns=new_col_names)

    data = data.fillna('NaN')
    data = data.dropna(how='any')

    return_list = []
    for index, row in data.iterrows():

        date_time = row['A']

        if type(date_time) == float:
            # dt = datetime.fromtimestamp(date_time)
            # date_time = dt.strftime("[%Y-%m-%d %H:%M:%S]")
            # date_time = dt.strftime("[%Y-%d-%m %H:%M:%S]")
            # date_time = str(date_time)
            dt = xlrd.xldate_as_datetime(date_time, 0)
            date_time = dt.strftime('%Y-%m-%d %H:%M:%S')

        # print(date_time)
        # print('date: ',date_time, type(date_time), 'in ', file_name)
        if type(date_time) == str:
            match = re.search(r'[0-9]+-[0-9]+-[0-9]+', date_time)
        elif type(date_time) == datetime:
            match = re.search(r'[0-9]+-[0-9]+-[0-9]+', date_time.strftime('%Y-%d-%m %H:%M:%S'))

        if match is None:
            continue

        # dt_ls = date_time.split('-')
        # if len(dt_ls[0]) == 4:
        #     dt_format = '%Y-%d-%m %H:%M:%S'
        # elif len(dt_ls[0]) == 2:
        #     dt_format = '%d-%m-%Y %H:%M:%S'
        # else:
        #     raise ValueError("Date is Invalid")

        # date_time_obj = datetime.strptime(date_time, dt_format)
        date_time_obj = check_time_format(str(date_time))
        # print(date_time_obj)

        pi = row['B']
        pd = row['C']
        ti = row['D']
        tm = row['E']
        vx = row['F']
        vz = row['G']
        vt = row['H']
        ct = row['I']
        lst = [date_time_obj, pi, pd, ti, tm, vx, vz, vt, ct]
        return_list.append(lst)

    return return_list


def sort_by_time(content_dictionary):
    for keys, value in content_dictionary.items():
        if value is not None:
            # print(keys)
            sorted_list = sorted(value, key=lambda x: x[0])
            content_dictionary[keys] = sorted_list
    return content_dictionary


def check_duplicates(content_dictionary):
    return_dictionary = {}
    duplicate_dates = {}
    bad_files = []
    for keys, value in content_dictionary.items():
        if value is not None:
            indices_to_delete = []
            break_flag = False
            # list_length = len(value[0])
            mod_list = value
            for i in range(0, len(mod_list) - 10, 10):
                for j in range(i + 1, i + 10):
                    # print(i, j)
                    # for k in range(list_length):
                    # try:

                    if mod_list[i][0] != mod_list[j][0]:
                        continue
                    elif mod_list[i][0] == mod_list[j][0]:
                        # found two similar date-times
                        bad_files.append(keys)

                        list1 = [float(x.replace(',', '')) if isinstance(x, str) else x for x in mod_list[i][1:]]
                        list2 = [float(x.replace(',', '')) if isinstance(x, str) else x for x in mod_list[j][1:]]

                        if list1 == list2:
                            # found two identical rows, delete two and keep one
                            # del mod_list[j]
                            indices_to_delete.append(j)
                        else:
                            # found two similar date-times with different parameters
                            # add file_name to a list
                            if keys in duplicate_dates.keys():
                                duplicate_dates[keys].append(mod_list[j])
                            else:
                                duplicate_dates[keys] = [mod_list[j]]
            #         except IndexError:
            #             break_flag = True
            #             break
            #     if break_flag:
            #         break
            # if break_flag:
            #     bad_files.append(keys)
            #     continue
            # else:
            my_list = [element for index, element in enumerate(mod_list) if index not in indices_to_delete]
            # print(my_list)
            content_dictionary[keys] = my_list
            # return_dictionary[keys] = my_list

    return content_dictionary, duplicate_dates, bad_files


def check_time_format(value):
    if type(value) == str:
        dt_ls = value.split('-')
        if len(dt_ls[0]) == 4:
            dt_format = '%Y-%d-%m %H:%M:%S'
        elif len(dt_ls[0]) == 2:
            dt_format = '%d-%m-%Y %H:%M:%S'
        else:
            raise ValueError("Date is Invalid: ", value)
            # print(value)
            # value = pd.to_datetime(value, origin='1900-01-01', unit='D')
            # dt_format = '%Y-%d-%m %H:%M:%S'
        value = datetime.strptime(value, dt_format)
        return value
    elif type(value) == datetime:
        return value

    # return value


def time_difference(content_dictionary):
    # files_checked = []
    for keys, value in content_dictionary.items():
        if value is None:
            continue

        temp_list = value
        for i in range(len(value) - 1):

            if type(value[i][1]) == str:
                continue

            try:
                value[i][0] = check_time_format(value[i][0])
            except ValueError:
                continue

            try:
                value[i + 1][0] = check_time_format(value[i + 1][0])
            except ValueError:
                continue

            time_diff = value[i + 1][0] - value[i][0]

            if time_diff.total_seconds() > 120 * 60:
                temp_list.insert(i + 1, ['GAP', 'From', value[i][0], 'To', value[i + 1][0], '-', '-', '-', '-'])
            elif time_diff.total_seconds() < 1:
                continue

            content_dictionary[keys] = temp_list
            # files_checked.append(keys)
    # print(files_checked)

    return content_dictionary




def insert_failures(content_dictionary, file_name):
    failures = pd.read_excel(file_name, names=['A', 'B', 'C'])
    for index, row in failures.iterrows():
        date = row['A']
        well = row['B']
        reason = row['C']
        key = well

        date = datetime.strptime(str(date), '%Y-%m-%d %H:%M:%S')
        row_list = [date, well, reason, '_', '_', 'FAILURE', 'FROM', 'THIS DATE']
        try:
            lst = content_dictionary[key]
            if lst is not None:
                for k in range(len(lst)):
                    if type(lst[k][0]) == datetime:
                        if date < lst[k][0]:
                            lst.insert(k, row_list)
                            content_dictionary[key] = lst
                            break
            else:
                none_files.append(key)
        except KeyError:
            print("KEY IN FAILURES EXCEL DOESN'T EXIST: ", key)
    return content_dictionary


def plot_timeline(content_dictionary):
    import matplotlib.pyplot as plt
    # import matplotlib.dates as mdates

    for key, value in content_dictionary.items():
        datetime_list = content_dictionary[key]
        if value is not None:

            failure_dates = []
            time_plot = [datetime_list[0][0]]
            for i in range(len(datetime_list)):
                if datetime_list[i][0] == 'GAP':
                    time_plot.append(datetime_list[i - 1][0])
                    time_plot.append(datetime_list[i + 1][0])
                elif datetime_list[i][-1] == 'THIS DATE':
                    failure_dates.append(datetime_list[i])

            time_plot.append(datetime_list[-1][0])
            datetime_ranges = [time_plot[i:i + 2] for i in range(0, len(time_plot), 2)]

            # Plotting the datetime range
            fig, ax = plt.subplots(figsize=(len(datetime_ranges) + 20, 4))

            mstart = 1
            mend = 2

            for i in range(len(datetime_ranges)):
                start = datetime_ranges[i][0]
                end = datetime_ranges[i][1]

                mstart += 2
                mend += 2

                ax.plot([mstart, mend], [0, 0], color='blue', linewidth=10)
                ax.plot([mend, mend + 1], [0, 0], color='red', linewidth=10)
                ax.text(mstart, 0.4, start, ha='center', va='bottom', rotation=90)
                ax.text(mend, 0.4, end, ha='center', va='bottom', rotation=90)

                for j in range(len(failure_dates)):
                    if start <= failure_dates[j][0] <= end:
                        print(failure_dates[j][2])
                        ax.scatter(mstart, -0.2, color='orange')
                        ax.text(mstart, -0.4, failure_dates[j][2], ha='center', va='bottom', rotation=0)

            ax.set_ylim(-0.7, 2)
            ax.set_yticks([])
            ax.set_xticks([])
            # ax.set_xlim(0, mend+5)
            # x_min = datetime_ranges[0][0] - timedelta(days=200)
            # x_max = datetime_ranges[-1][1] + timedelta(days=200)
            # ax.set_xscale('log')
            # ax.set_xlim(x_min, x_max)
            ax.set_xlabel('Datetime')
            ax.set_title(key)
            # ax.plot(datetime_ranges[0][0], datetime_ranges[-1][-1], [1,1], color='red', linewidth=2)

            # date_fmt = mdates.DateFormatter('%Y-%m-%d %H:%M:%S')
            # ax.xaxis.set_major_formatter(date_fmt)
            # ax.xaxis.set_major_locator(mdates.AutoDateLocator())
            # plt.xticks(rotation=90)
            print('working', key)
            plot_file_name = str(key) + '.png'
            plt.savefig(plot_file_name)

            plt.show()


if __name__ == '__main__':
    level1_dir = os.listdir(ROOT_PATH)
    # print(level1_dir)
    wrong_files = []
    for l1 in level1_dir:
        file_names = os.listdir(ROOT_PATH + l1)
        for file_n in file_names:
            file_path = ROOT_PATH + l1 + '/' + file_n
            print(file_path)
            field, well = get_field_well(file_n)
            # print(f'field = {field}  --- well name = {well} ------ top directory = {l1} --file={file_n} -- type: {type(field), type(well)}')

            # if field is None or well is None:
            if well is None:
                with open('vague_files.txt', 'a') as file:
                    str_write = "Field: " + str(field) + " , " + "Well: " + str(well) + "\n"
                    file.write(str_write)
                continue

            # print("INCLUDING FILE" + '\n')

            # key = field + "_" + well
            key = well
            # print(key)
            if key in content_dict.keys():
                contents = read_file_contents(file_path)
                if contents is not None:
                    try:
                        content_dict[key].extend(contents)
                    except ValueError:
                        wrong_files.append(file_path)
            else:
                contents = read_file_contents(file_path)
                if contents is not None:
                    try:
                        content_dict[key] = read_file_contents(file_path)
                    except ValueError:
                        wrong_files.append(file_path)
    pop_from_dict = []
    for key, value in content_dict.items():
        if value is None:
            none_files.append(key)
            pop_from_dict.append(key)

    for i in range(len(pop_from_dict)):
        content_dict.pop(pop_from_dict[i])

    from datetime import timedelta
    for key, value in content_dict.items():
        if len(value) > 100:
            delta = value[20][0] - value[19][0]
            if delta == timedelta(seconds=300):
                temp_list = value[::2]
            elif delta == timedelta(seconds=600):
                temp_list = value
            elif delta == timedelta(seconds=30):
                temp_list = value[::20]
            elif delta == timedelta(seconds=60):
                temp_list = value[::10]
            else:
                continue
            content_dict[key] = temp_list
        del delta
        del temp_list

    content_dict_sorted_time = sort_by_time(content_dict)
    content_dict_unique, files_with_duplicates, bad_files = check_duplicates(content_dict_sorted_time)
    content_dict_gaps = time_difference(content_dict_unique)
    content_dict_f = insert_failures(content_dict_gaps, 'failures.xlsx')

    # content_dict_aa = insert_failures(content_dict, 'failures.xlsx')
    # content_dict_sorted_time = sort_by_time(content_dict_aa)
    # content_dict_unique, files_with_duplicates, bad_files = check_duplicates(content_dict_sorted_time)
    # content_dict_f = time_difference(content_dict_unique)

    # content_dict, files_with_duplicates, bad_files = check_duplicates(content_dict)

    # print(content_dict)

    # print("FILES NOT WORKING WITH DUP FUNCTION", bad_files)

    # OUTPUT TEXT FILES
    # for key, value in content_dict_f.items():
    #     text_file = key + '.txt'
    #     with open(text_file, 'w') as f:
    #         for i in range(len(value)):
    #             f.write(str(content_dict_f[key][i]) + '\n')

    import openpyxl
    from openpyxl.styles import PatternFill

    none_values_names = []

    for key, value in content_dict_f.items():

        if value is None:
            none_values_names.append(key)
            continue

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        filename = key + '.xlsx'

        break_flag = False

        for i in range(len(value)):
            try:
                sheet.append(value[i])
                if value[i][0] == 'GAP':
                    for cell in sheet[i + 1]:
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            except ValueError:
                # print("PROBLEM WITH EXCEL: ", key, "--Value Type: ", type(value))
                break_flag = True
                break
            if break_flag:
                continue

        workbook.save(filename)

    plot_timeline(content_dict_f)
    # print(files_w_one_column)

    with open('files_with_no_pd.txt', 'w') as file:
        for row in files_w_no_pd:
            file.write(row + '\n')

    bad_files_unique = set(bad_files)
    with open('files_with_duplicate_dates.txt', 'w') as file:
        for wellnames in bad_files_unique:
            file.write(str(wellnames) + '\n')

    with open('files_with_duplicate_dates_different_values.txt', 'w') as file:
        for well, dates in files_with_duplicates.items():
            file.write(str(well)+ '\n')

    with open('files_with_duplicates.txt', 'w') as file:
        for well, dates in files_with_duplicates.items():
            file.write(str(well)+ ' ----> '+ str(dates) + '\n')
            file.write('\n')
            file.write('\n')



    bad_files_temp = bad_files_unique
    files_with_duplicates_set = set(files_with_duplicates.keys())

    bad_files_temp.difference_update(files_with_duplicates_set)

    # for item in bad_files_unique:
    #     if item in files_with_duplicates_set:
    #         bad_files_temp.remove(item)

    with open('files_with_duplicate_dates_same_values.txt', 'w') as file:
        for wellnames in bad_files_temp:
            file.write(str(wellnames) + '\n')




print(none_files)
